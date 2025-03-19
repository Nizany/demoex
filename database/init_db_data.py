from openpyxl import load_workbook

from database.connection import Base, engine, session
from database.models.Material import Material
from database.models.Order import Order
from database.models.Partner import Partner
from database.models.Product import Product
from database.models.ProductType import ProductType

# база данных создается
Base.metadata.drop_all(engine)
Base.metadata.create_all(engine)

product_type_file = "excel/Product_type_import.xlsx"
products_file = "excel/Products_import.xlsx"
partners_file = "excel/Partners_import.xlsx"
sales_history_file = "excel/Partner_products_import.xlsx"
materials_file = "excel/Material_type_import.xlsx"


def import_product_types(session):
    wb = load_workbook(product_type_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        type, coefficient_of_product_type = row
        if not type or not coefficient_of_product_type:
            continue
        session.add(ProductType(type=type, coefficient_of_product_type=coefficient_of_product_type))
    session.commit()
    print("Типы продукции добавлены.")


def import_products(session):
    wb = load_workbook(products_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        fk_type, name, article, min_cost = row
        product_type = session.query(ProductType).filter_by(type=fk_type).first()
        session.add(Product(
            article=article,
            name=name,
            min_cost=min_cost,
            fk_type=product_type.id
        ))
    session.commit()
    print("Продукты добавлены.")


def import_partners(session):
    wb = load_workbook(partners_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        type, company_name, boss_name, mail, phone_number, address, inn, rank = row
        session.add(Partner(
            type=type,
            company_name=company_name,
            address=address,
            inn=inn,
            boss_name=boss_name,
            phone_number=phone_number,
            mail=mail,
            rank=rank
        ))
    session.commit()
    print("Партнёры добавлены.")


def import_sales_history(session):
    wb = load_workbook(sales_history_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        product_name, partner_name, amount_of_products, date_of_create = row
        product = session.query(Product).filter_by(name=product_name).first()
        partner = session.query(Partner).filter_by(company_name=partner_name).first()
        session.add(Order(
            fk_product_id=product.id,
            fk_company_id=partner.id,
            amount_of_products=amount_of_products,
            date_of_create=date_of_create
        ))
    session.commit()
    print("История продаж добавлена.")


def import_materials(session):
    wb = load_workbook(materials_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        type, percentage_of_defective_material = row
        session.add(Material(
            type=type,
            percentage_of_defective_material=percentage_of_defective_material
        ))
    session.commit()
    print("Материалы добавлены.")


try:
    import_product_types(session)
    import_products(session)
    import_partners(session)
    import_sales_history(session)
    import_materials(session)
except Exception as e:
    session.rollback()
    print(f"Ошибка при импорте данных: {e}")
finally:
    session.close()
    print("Импорт данных завершён.")